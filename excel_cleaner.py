import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ExcelColumnClearer:
    def __init__(self):
        self.folder_path = ""
        self.column_number = 0

    def set_folder_path(self, folder_path):
        self.folder_path = folder_path

    def set_column_number(self, column_number):
        self.column_number = column_number

    def clear_column_in_files(self):
        if not self.folder_path or self.column_number <= 0:
            raise ValueError("请先设置有效的文件夹路径和列号")

        processed_files = 0

        for root, dirs, files in os.walk(self.folder_path):
            for file in files:
                if file.endswith(('.xlsx', '.xls')):
                    file_path = os.path.join(root, file)
                    try:
                        # 使用 openpyxl 加载工作簿以保留格式
                        wb = load_workbook(file_path)
                        ws = wb.active
                        
                        # 获取要清空的列的字母索引
                        col_letter = get_column_letter(self.column_number)
                        
                        # 清空指定列的内容（跳过表头）
                        for row in range(2, ws.max_row + 1):
                            ws[f'{col_letter}{row}'].value = None
                            
                        # 保存文件
                        wb.save(file_path)
                        processed_files += 1

                    except Exception as e:
                        print(f"处理文件 {file} 时出错：{str(e)}")
                        continue

        return processed_files